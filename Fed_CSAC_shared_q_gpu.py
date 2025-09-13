import numpy as np
import matplotlib.pyplot as plt
import torch
import torch.nn as nn
import pandas as pd
import os
from env import CombinedEnergyEnv
from env import da_market_clearing, get_market_prices_car
from single_ies_shared_q_gpu import C_SAC_GPU
from env.carbon import calculate_carbon_quota_split

# 获取时序定价
tou_buy, fit_sell, car_buy, car_sell, grid_co2 = get_market_prices_car()

def federated_weighted_average_gpu(agents, personal_steps=1):
    """
    PyTorch版本的等权联邦聚合 + 个性化更新
    - 等权聚合共享浅层（Actor_Shared_*, Critic_Shared_*）
    - 在本函数内对个性化深层做本地更新：临时冻结共享层，调用现有 ag.replay() / ag._update_from_batch()
    """

    # --------- 小工具：按前缀冻结/解冻参数 ----------
    def set_trainable_by_prefix(model, prefixes, trainable):
        if isinstance(prefixes, str):
            prefixes = (prefixes,)
        for name, param in model.named_parameters():
            if any(name.startswith(p) for p in prefixes):
                param.requires_grad = trainable

    # --------- 小工具：构建"等权聚合"的参数名->权重字典（仅匹配指定前缀） ----------
    def build_equal_avg_weights_gpu(models, prefixes):
        if isinstance(prefixes, str):
            prefixes = (prefixes,)
        
        # 收集所有匹配的参数名
        param_names = set()
        for model in models:
            for name in model.state_dict().keys():
                # 将PyTorch参数名格式适配到前缀匹配
                formatted_name = name.replace(".", "_")
                if any(formatted_name.startswith(p) for p in prefixes):
                    param_names.add(name)
        
        # 逐参数做等权平均
        name_to_avg = {}
        for param_name in param_names:
            param_tensors = []
            for model in models:
                if param_name in model.state_dict():
                    param_tensors.append(model.state_dict()[param_name])
            
            if param_tensors:
                # 计算平均值
                avg_param = torch.mean(torch.stack(param_tensors), dim=0)
                name_to_avg[param_name] = avg_param
                
        return name_to_avg

    # --------- 1) 先做"个性化深层"的本地更新（冻结共享层，只训深层） ----------
    # 共享层前缀（适配PyTorch命名约定）
    ACTOR_SHARED_PREFIX = "shared_layers"  # PyTorch中共享层的命名
    CRITIC_SHARED_PREFIXES = ("shared_layers","ln_first")

    for ag in agents:
        # 冻结共享层
        set_trainable_by_prefix(ag.actor, ACTOR_SHARED_PREFIX, False)
        set_trainable_by_prefix(ag.critic, CRITIC_SHARED_PREFIXES, False)

        # 个性化更新若干步（只会更新未被冻结的"深层"）
        for _ in range(max(1, personal_steps)):
            batch = ag.replay()
            if batch is not None:
                ag._update_from_batch(batch)

        # 恢复共享层可训练标志（不改变外部训练行为）
        set_trainable_by_prefix(ag.actor, ACTOR_SHARED_PREFIX, True)
        set_trainable_by_prefix(ag.critic, CRITIC_SHARED_PREFIXES, True)

    # --------- 2) 对"共享浅层"做等权聚合并下发 ----------
    # Actor 共享浅层
    actor_models = [ag.actor for ag in agents]
    actor_avg = build_equal_avg_weights_gpu(actor_models, prefixes=ACTOR_SHARED_PREFIX)

    # Critic 共享浅层（单模型双头：只聚合主干浅层）
    critic_models = [ag.critic for ag in agents]
    critic_avg = build_equal_avg_weights_gpu(critic_models, prefixes=CRITIC_SHARED_PREFIXES)

    # 应用平均权重到所有 agent 的共享层（个性化层保持本地参数）
    def apply_avg_gpu(model, avg_dict):
        with torch.no_grad():
            state_dict = model.state_dict()
            for param_name, avg_weight in avg_dict.items():
                if param_name in state_dict:
                    state_dict[param_name].copy_(avg_weight)

    for ag in agents:
        apply_avg_gpu(ag.actor, actor_avg)
        apply_avg_gpu(ag.critic, critic_avg)
        # 目标网络软更新保持不变
        ag.soft_update_all_targets()

# def federated_weighted_average_gpu(agents, personal_steps=1, verbose=True):
#     """
#     PyTorch版本的等权联邦聚合 + 个性化更新
#     - 等权聚合共享浅层（Actor_Shared_*, Critic_Shared_*）
#     - 在本函数内对个性化深层做本地更新：临时冻结共享层，调用现有 ag.replay() / ag._update_from_batch()
#     - verbose=True 时打印聚合的详细信息用于验证
#     """

#     # --------- utils: prefix freeze/unfreeze ----------
#     def set_trainable_by_prefix(model, prefixes, trainable):
#         if isinstance(prefixes, str):
#             prefixes = (prefixes,)
#         for name, param in model.named_parameters():
#             if any(name.startswith(p) for p in prefixes):
#                 param.requires_grad = trainable

#     # --------- utils: avg over prefixes ----------
#     def build_equal_avg_weights_gpu(models, prefixes):
#         if isinstance(prefixes, str):
#             prefixes = (prefixes,)
#         param_names = set()
#         for model in models:
#             for name in model.state_dict().keys():
#                 formatted_name = name.replace(".", "_")
#                 if any(formatted_name.startswith(p) for p in prefixes):
#                     param_names.add(name)
#         name_to_avg = {}
#         for param_name in param_names:
#             tensors = [m.state_dict()[param_name] for m in models if param_name in m.state_dict()]
#             if tensors:
#                 name_to_avg[param_name] = torch.mean(torch.stack(tensors), dim=0)
#         return name_to_avg

#     # --------- helpers for printing/verification ----------
#     def _keys_with_prefix(sd, prefixes):
#         if isinstance(prefixes, str):
#             prefixes = (prefixes,)
#         return sorted([k for k in sd.keys() if any(k.startswith(p) for p in prefixes)])

#     @torch.no_grad()
#     def _max_pairwise_diff_across_agents(param_name, models):
#         vals = [m.state_dict()[param_name] for m in models]
#         maxd = 0.0
#         for i in range(len(vals)):
#             for j in range(i+1, len(vals)):
#                 d = (vals[i] - vals[j]).abs().max().item()
#                 if d > maxd: maxd = d
#         return maxd

#     # --------- 1) personalization with shared frozen ----------
#     ACTOR_SHARED_PREFIX = "shared_layers"
#     CRITIC_SHARED_PREFIXES = ("shared_layers", "ln_first")

#     if verbose:
#         print("[Fed] Personalization phase: freezing shared prefixes:",
#               {"actor": ACTOR_SHARED_PREFIX, "critic": list(CRITIC_SHARED_PREFIXES)})

#     for ag in agents:
#         set_trainable_by_prefix(ag.actor,  ACTOR_SHARED_PREFIX,     False)
#         set_trainable_by_prefix(ag.critic, CRITIC_SHARED_PREFIXES,  False)

#         for _ in range(max(1, personal_steps)):
#             batch = ag.replay()
#             if batch is not None:
#                 ag._update_from_batch(batch)

#         set_trainable_by_prefix(ag.actor,  ACTOR_SHARED_PREFIX,     True)
#         set_trainable_by_prefix(ag.critic, CRITIC_SHARED_PREFIXES,  True)

#     # --------- 2) averaging only shared ----------
#     actor_models  = [ag.actor  for ag in agents]
#     critic_models = [ag.critic for ag in agents]

#     actor_avg  = build_equal_avg_weights_gpu(actor_models,  prefixes=ACTOR_SHARED_PREFIX)
#     critic_avg = build_equal_avg_weights_gpu(critic_models, prefixes=CRITIC_SHARED_PREFIXES)

#     if verbose:
#         # 打印将被聚合的键数量与示例
#         a_keys = sorted(actor_avg.keys())
#         c_keys = sorted(critic_avg.keys())
#         print(f"[Fed] Actor shared keys to avg: {len(a_keys)}",
#               (a_keys[:5] + ["..."] if len(a_keys) > 5 else a_keys))
#         print(f"[Fed] Critic shared keys to avg: {len(c_keys)}",
#               (c_keys[:5] + ["..."] if len(c_keys) > 5 else c_keys))

#         # 聚合前：检查共享键在各 agent 间的最大两两差（越小越说明个性化阶段确实冻结了共享层）
#         if a_keys:
#             maxd_before_actor  = max(_max_pairwise_diff_across_agents(k, actor_models)  for k in a_keys)
#             print(f"[Fed] Max pairwise diff across agents BEFORE avg (Actor shared):  {maxd_before_actor:.3e}")
#         if c_keys:
#             maxd_before_critic = max(_max_pairwise_diff_across_agents(k, critic_models) for k in c_keys)
#             print(f"[Fed] Max pairwise diff across agents BEFORE avg (Critic shared): {maxd_before_critic:.3e}")

#         # 专门点名 ln_first
#         ln_keys = []
#         if critic_models:
#             sample_sd = critic_models[0].state_dict()
#             ln_keys = _keys_with_prefix(sample_sd, "ln_first")
#             print(f"[Fed] Critic ln_first.* keys count: {len(ln_keys)}",
#                   (ln_keys[:5] + ["..."] if len(ln_keys) > 5 else ln_keys))

#     # 覆写共享层平均权重
#     @torch.no_grad()
#     def apply_avg_gpu(model, avg_dict):
#         sd = model.state_dict()
#         for k, v in avg_dict.items():
#             if k in sd:
#                 sd[k].copy_(v)

#     for ag in agents:
#         apply_avg_gpu(ag.actor,  actor_avg)
#         apply_avg_gpu(ag.critic, critic_avg)
#         ag.soft_update_all_targets()  # keep as is

#     if verbose:
#         # 聚合后：共享层在各 agent 间应当几乎完全一致
#         if actor_avg:
#             maxd_after_actor  = max(_max_pairwise_diff_across_agents(k, actor_models)  for k in actor_avg.keys())
#             print(f"[Fed] Max pairwise diff across agents AFTER  avg (Actor shared):  {maxd_after_actor:.3e}")
#         if critic_avg:
#             maxd_after_critic = max(_max_pairwise_diff_across_agents(k, critic_models) for k in critic_avg.keys())
#             print(f"[Fed] Max pairwise diff across agents AFTER  avg (Critic shared): {maxd_after_critic:.3e}")

#         # 快速抽样检查：个性化层仍保留差异（挑几条非共享键看差异是否>0）
#         def any_personal_diff(models, shared_prefixes):
#             if isinstance(shared_prefixes, str):
#                 shared_prefixes = (shared_prefixes,)
#             sample_sd = models[0].state_dict()
#             for k in sample_sd.keys():
#                 if any(k.startswith(p) for p in shared_prefixes):
#                     continue
#                 # 跳过 buffers 的统计项（一般都是参数，这里简单处理）
#                 try:
#                     d = _max_pairwise_diff_across_agents(k, models)
#                 except:
#                     continue
#                 if d > 0:
#                     return True
#             return False

#         print(f"[Fed] Actor personal params keep diversity?  {any_personal_diff(actor_models,  ACTOR_SHARED_PREFIX)}")
#         print(f"[Fed] Critic personal params keep diversity? {any_personal_diff(critic_models, CRITIC_SHARED_PREFIXES)}")


def compute_c_n(p_grid_trade, car_emis, t, MG_car=8800):
    if p_grid_trade > 0:
        C_n = grid_co2[t] * p_grid_trade + car_emis - MG_car  # 8800占位，不同MG8800位置不一样
    else:
        C_n = car_emis - MG_car
    return C_n


def train_multi_agents_gpu(
    scenarios,
    max_rounds=5000,
    max_steps=24,
    agg_interval=1,
    gamma=0.99,
    tau=0.005,
    device=None,
    log_dir="step_logs",          # 新增：日志目录
    log_flush_interval=1,          # 新增：每多少轮写一次盘（1 表示每轮写）
):
    """
    PyTorch版本的多 MG 串并联训练：
    每轮:
      每步: 所有 MG act-> step -> 集中撮合 -> compute_trade_cost -> remember -> update
      每 agg_interval 轮后 FedAvg
    """

        # ---------------- 工具函数：通用数值/向量安全转化 ----------------
    def _to_float(x):
        # 标量 -> float
        if x is None:
            return float("nan")
        if isinstance(x, (float, int)):
            return float(x)
        if isinstance(x, (np.floating, np.integer)):
            return float(x.item())
        if torch.is_tensor(x):
            if x.numel() == 1:
                return float(x.detach().cpu().item())
            else:
                return float("nan")
        try:
            return float(x)
        except Exception:
            return float("nan")

    def _to_list(x):
        # 张量/ndarray/可迭代 -> Python list（用于保存到 Excel 的字符串列）
        if x is None:
            return None
        if torch.is_tensor(x):
            return x.detach().cpu().numpy().tolist()
        if isinstance(x, np.ndarray):
            return x.tolist()
        if isinstance(x, (list, tuple)):
            return list(x)
        # 标量也收敛成单元素 list 便于查看
        if isinstance(x, (float, int, np.floating, np.integer)):
            return [float(x)]
        return [str(x)]
    
    # 设备设置
    if device is None:
        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Training on device: {device}")

    # 初始化 agents
    agents = []

    for s in scenarios:
        env = CombinedEnergyEnv(s)
        ag = C_SAC_GPU(env, gamma=gamma, tau=tau, device=device)
        agents.append(ag)
    
    n_agents = len(agents)
    critic_loss_hist   = [ [] for _ in range(n_agents) ]
    actor_loss_hist    = [ [] for _ in range(n_agents) ]
    qc_loss_hist       = [ [] for _ in range(n_agents) ]
    reward_hist        = [ [] for _ in range(n_agents) ]
    constraint_hist    = [ [] for _ in range(n_agents) ]

    critic_loss_ep_hist = [[] for _ in range(n_agents)]
    actor_loss_ep_hist  = [[] for _ in range(n_agents)]
    qc_loss_ep_hist     = [[] for _ in range(n_agents)]

        # ---------------- 日志初始化（Excel）----------------
    os.makedirs(log_dir, exist_ok=True)
    agent_excel_paths = [os.path.join(log_dir, f"agent{idx}_step_logs.xlsx") for idx in range(n_agents)]
    step_logs_buffer = [[] for _ in range(n_agents)]  # 每个 agent 一份 list[dict]

    def append_df_to_excel(path, df, sheet_name="logs"):
        """把 DataFrame 追加到 Excel 的某个 sheet（不存在则创建，并写表头）"""
        from openpyxl import load_workbook

        if not os.path.exists(path):
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            return

        # 文件已存在：定位追加起始行
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            startrow = wb[sheet_name].max_row  # 包含表头
        else:
            startrow = 0

        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=(startrow == 0),
                startrow=startrow
            )

    def _flush_round_logs(rnd):
        """把当前轮缓存写入各自的 Excel。"""
        nonlocal step_logs_buffer
        for idx in range(n_agents):
            rows = step_logs_buffer[idx]
            if not rows:
                continue
            df = pd.DataFrame(rows)
            append_df_to_excel(agent_excel_paths[idx], df, sheet_name="logs")
            step_logs_buffer[idx] = []  # 清空缓存

    # 全局训练轮
    for rnd in range(1, max_rounds+1):
        print(f"== 轮次 {rnd}/{max_rounds} ==")
        # 重置 envs
        states = [ag.env.reset() for ag in agents]
        dones  = [False]*len(agents)
        step   = 0

        # 每个 agent 本轮的累积指标
        total_reward_agent  = [0.0]*n_agents
        total_penalty_agent = [0.0]*n_agents

        # 单个 episode
        while not all(dones) and step < max_steps:
            # 并行选动作
            actions = []
            samples = []  
            for ag, s in zip(agents, states):
                a, sampled = ag.act(s)
                actions.append(a)
                samples.append(sampled)

            # 并行执行 env.step，并收集 P_n/C_n 和本地 reward/emis
            local_emis = [0.0]*len(agents)
            local_cost_operate = []
            infos      = []
            P_buys, P_sells, C_buys, C_sells = [], [], [], []
            P_n_record = [0.0]*len(agents)
            C_n_record = [0.0]*len(agents)
            P_load_record = [0.0]*len(agents)

            pb_rec = [0.0]*n_agents; ps_rec = [0.0]*n_agents
            cb_rec = [0.0]*n_agents; cs_rec = [0.0]*n_agents

            for idx, (ag, a) in enumerate(zip(agents, actions)):
                # 只更新部分状态，获得 P_n 等
                _, done, info = ag.env.step(a)
                local_cost_operate.append(info['operate_cost'])
                infos.append(info)
                dones[idx] = done

                # 收集挂单信息
                P_n = info['P_n']
                P_n_record[idx] = P_n
                local_emis[idx] = info['carbon_emis']
                P_load_record[idx] = info['P_load']

                # 占位价格存在旧 state 的 8-11
                pb_rec[idx] = states[idx][8];  ps_rec[idx] = states[idx][9]
                cb_rec[idx] = states[idx][10]; cs_rec[idx] = states[idx][11]
                # 电碳盈余有正负
                if P_n > 0:  P_buys.append((idx, pb_rec[idx],  P_n))
                else:        P_sells.append((idx, ps_rec[idx], -P_n))

            # 市场撮合
            # 电
            lambda_e_buy  = tou_buy[step]
            ele_res, e_p_m_buy, e_p_m_sell, e_summary = da_market_clearing(
                P_buys, P_sells,
                lambda_buy=lambda_e_buy, lambda_sell=fit_sell)
            
            print(f'ele_res:{ele_res}')
            
            for i, v in ele_res.items():
                p_grid_trade = v['grid_qty']
                car_emis = local_emis[i]
                MG_car = calculate_carbon_quota_split(P_load_record[i]).get('quota_total', 37000)
                C_n = compute_c_n(p_grid_trade, car_emis, step, MG_car=MG_car)
                C_n_record[i] = C_n

                if C_n > 0: C_buys.append((i, cb_rec[i],  C_n))
                else:       C_sells.append((i, cs_rec[i],  -C_n))

            # 碳
            car_res, car_p_m_buy, car_p_m_sell, car_summary = da_market_clearing(
                C_buys, C_sells,
                lambda_buy=car_buy, lambda_sell=car_sell)

            print(f'car_res:{car_res}')
            # 执行交易成本计算及完整状态更新，并存储/训练
            next_states = []
            for idx, ag in enumerate(agents):
                info = infos[idx]
                # compute_trade_cost 返回 (trade_cost, full_state)
                trade_cost, full_state = ag.env.compute_trade_cost(
                    P_n_record[idx], C_n_record[idx],
                    elec_price_buy  = e_p_m_buy,
                    elec_price_sell = e_p_m_sell,
                    carbon_price_buy  = car_p_m_buy,
                    carbon_price_sell = car_p_m_sell
                )
                # 组合总 reward
                new_r = -(
                    local_cost_operate[idx] + trade_cost
                )

                total_reward_agent[idx]  += new_r*1e-6
                total_penalty_agent[idx] += info['penalty']*1e-8

                # 记忆使用完整状态
                ag.remember(
                    states[idx], samples[idx],
                    new_r, full_state, dones[idx], info['penalty']
                )
                # 更新网络
                q1q2 = float("nan")
                qc   = float("nan")
                al   = float("nan")
                batch = ag.replay()
                if batch is not None:
                    q1_l, q2_l, qc_l, a_l = ag._update_from_batch(batch)
                    critic_loss_hist[idx].append( (q1_l + q2_l) / 2 )
                    qc_loss_hist[idx].append( qc_l )
                    actor_loss_hist[idx].append( a_l )
                    q1q2 = _to_float((q1_l + q2_l) / 2)
                    qc = _to_float(qc_l)
                    al = _to_float(a_l)

                # ---- 逐步日志：为当前 agent/step 记录一行（将写入 Excel）----
                row = {
                    "round": rnd,
                    "step": step,
                    "agent": idx,
                    # 状态与动作
                    "state": _to_list(states[idx]),
                    "action": _to_list(actions[idx]),
                    "sampled": _to_list(samples[idx]),
                    "next_state": _to_list(full_state),
                    "done": bool(dones[idx]),
                    # 本地量
                    "operate_cost": _to_float(local_cost_operate[idx]),
                    "penalty": _to_float(info.get("penalty", float("nan"))),
                    "carbon_emis": _to_float(info.get("carbon_emis", float("nan"))),
                    "P_load": _to_float(P_load_record[idx]),
                    # 撮合输入（挂单）
                    "P_n": _to_float(P_n_record[idx]),
                    "C_n": _to_float(C_n_record[idx]),
                    # "pb_quote": _to_float(pb_rec[idx]),
                    # "ps_quote": _to_float(ps_rec[idx]),
                    # "cb_quote": _to_float(cb_rec[idx]),
                    # "cs_quote": _to_float(cs_rec[idx]),
                    # 撮合结果关键信息（统一价；grid 交易量）
                    "e_clearing_buy": _to_float(e_p_m_buy),
                    "e_clearing_sell": _to_float(e_p_m_sell),
                    "c_clearing_buy": _to_float(car_p_m_buy),
                    "c_clearing_sell": _to_float(car_p_m_sell),
                    "grid_qty_e": _to_float(ele_res.get(idx, {}).get("grid_qty", 0.0)),
                    "grid_qty_c": _to_float(car_res.get(idx, {}).get("grid_qty", 0.0)),
                    # 经济量
                    "trade_cost": _to_float(trade_cost),
                    "reward": _to_float(new_r),
                    # 学习损失
                    "critic_loss": _to_float(q1q2),
                    "qc_loss": _to_float(qc),
                    "actor_loss": _to_float(al),
                }
                step_logs_buffer[idx].append(row)

                next_states.append(full_state)

            # 迭代
            states = next_states
            step  += 1

        # Episode 结束：打印并记录每个 agent 的指标
        for ies in range(n_agents):
            print(f" Agent{ies} | Reward: {total_reward_agent[ies]:.3f} | Penalty: {total_penalty_agent[ies]:.3f}")
            reward_hist[ies].append(total_reward_agent[ies])
            constraint_hist[ies].append(total_penalty_agent[ies])

            if critic_loss_hist[ies]:
                critic_loss_ep_hist[ies].append(critic_loss_hist[ies][-1])
            if actor_loss_hist[ies]:
                actor_loss_ep_hist[ies].append(actor_loss_hist[ies][-1])
            if qc_loss_hist[ies]:
                qc_loss_ep_hist[ies].append(qc_loss_hist[ies][-1])

        # 每 log_flush_interval 轮写盘一次（Excel 追加）
        if rnd % log_flush_interval == 0:
            _flush_round_logs(rnd)

        # 周期联邦平均
        if rnd % agg_interval == 0:
            print(f"-- 执行联邦聚合 (轮次 {rnd}) --")
            federated_weighted_average_gpu(agents)

    # 训练循环结束，如仍有未写缓存，最后再 flush
    _flush_round_logs(max_rounds)
    
    # 训练结束后画图 —— 每个变量单独一张图


    # 确保输出目录存在
    outdir = "image_result"
    os.makedirs(outdir, exist_ok=True)

    def draw_and_save(y, title, ylabel, fname, label=None, xlabel=None):
        if not y:  # 空数据直接跳过
            return
        plt.figure()
        if label is None:
            plt.plot(y)
        else:
            plt.plot(y, label=label)
            plt.legend()
        if xlabel: plt.xlabel(xlabel)
        plt.ylabel(ylabel)
        plt.title(title)
        plt.grid(True)
        plt.tight_layout()
        # 保存 PNG
        save_path = os.path.join(outdir, fname)
        plt.savefig(save_path, dpi=150, bbox_inches="tight")
        plt.show()   # 需要同时显示就保留；若不显示可改为 plt.close()
        plt.close()  # 如果不想弹图窗口，把上面的 plt.show() 注释掉，并打开这一行

    # 训练结束后画图 —— 每个变量单独一张图，并保存
    for idx in range(n_agents):
        draw_and_save(
            reward_hist[idx],
            title=f"Agent {idx} - Reward per Episode",
            ylabel="Reward",
            fname=f"agent{idx}_reward_gpu.png",
            label="Reward",
            xlabel="Episode"
        )
        draw_and_save(
            constraint_hist[idx],
            title=f"Agent {idx} - Penalty per Episode",
            ylabel="Penalty",
            fname=f"agent{idx}_penalty_gpu.png",
            label="Penalty",
            xlabel="Episode"
        )

    for idx in range(n_agents):
        draw_and_save(
            critic_loss_ep_hist[idx],
            title=f"Agent {idx} - Critic Loss",
            ylabel="Loss",
            fname=f"agent{idx}_critic_loss_gpu.png",
            label="Critic Loss",
            xlabel="Episode"
        )
        draw_and_save(
            actor_loss_ep_hist[idx],
            title=f"Agent {idx} - Actor Loss",
            ylabel="Loss",
            fname=f"agent{idx}_actor_loss_gpu.png",
            label="Actor Loss",
            xlabel="Episode"
        )
        draw_and_save(
            qc_loss_ep_hist[idx],
            title=f"Agent {idx} - Constraint(Qc) Loss",
            ylabel="Loss",
            fname=f"agent{idx}_qc_loss_gpu.png",
            label="Constraint Loss",
            xlabel="Episode"
        )

    return agents


def save_agents(agents, filepath):
    """保存所有智能体的模型"""
    saved_data = {}
    for i, agent in enumerate(agents):
        agent_data = {
            'actor_state_dict': agent.actor.state_dict(),
            'critic_state_dict': agent.critic.state_dict(),
            'qc_critic_state_dict': agent.qc_critic.state_dict(),
            'alpha': agent.alpha.data,
            'lambda_': agent.lambda_.data,
        }
        saved_data[f'agent_{i}'] = agent_data
    
    torch.save(saved_data, filepath)
    print(f"Models saved to {filepath}")


def load_agents(agents, filepath):
    """加载所有智能体的模型"""
    saved_data = torch.load(filepath, map_location='cpu')
    
    for i, agent in enumerate(agents):
        agent_key = f'agent_{i}'
        if agent_key in saved_data:
            agent_data = saved_data[agent_key]
            agent.actor.load_state_dict(agent_data['actor_state_dict'])
            agent.critic.load_state_dict(agent_data['critic_state_dict'])
            agent.qc_critic.load_state_dict(agent_data['qc_critic_state_dict'])
            agent.alpha.data = agent_data['alpha']
            agent.lambda_.data = agent_data['lambda_']
    
    print(f"Models loaded from {filepath}")


if __name__ == '__main__':
    # 设备选择
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Using device: {device}")
    
    scenarios = ['IES1', 'IES2', 'IES3']
    # scenarios = ['IES1']  # 单智能体测试
    
    agents = train_multi_agents_gpu(
        scenarios, 
        max_rounds=1, 
        device=device
    )
    
    # 保存训练好的模型
    save_agents(agents, "federated_models_gpu.pth")
    
    print("Training completed successfully!")